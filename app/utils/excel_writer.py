from openpyxl import load_workbook
import shutil
from pathlib import Path

def generate_filled_excel(source_path: str, output_path: str, pending_items: list):
    """
    pending_items: dict のリスト
        {
            "excel_row": int,
            "qty": int
        }
    """
    # 原本コピー
    shutil.copyfile(source_path, output_path)

    # コピーしたファイルに書き込む
    wb = load_workbook(output_path)
    ws = wb.active
    
    if ws is None:
        raise ValueError("ワークシートが見つかりません")

    QTY_COL = "D"

    for item in pending_items:
        row = item["excel_row"]
        qty = item["qty"]
        # セルに値を設定（int型に変換して安全に書き込む）
        cell = ws[f"{QTY_COL}{row}"]
        cell.value = int(qty) if qty is not None else 0

    wb.save(output_path)
    return output_path
