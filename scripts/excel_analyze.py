# excel_inspector.py
import os
from pathlib import Path
import openpyxl

# ======== 設定項目 ========
TARGET_COLUMN = "C"
START_ROW = 5
END_ROW = 65
# ==========================


def classify_blank(value):
    """空白の種類を分類する"""
    if value is None:
        return "NONE (cell truly empty)"
    if isinstance(value, str):
        if value == "":
            return "EMPTY_STRING ('')"
        # 全角スペースや不可視文字を含めたstrip判定
        stripped = value.strip()
        if stripped == "":
            return f"WHITESPACE_ONLY (len={len(value)})"
    return None  # 空白ではない


def find_excel_file() -> Path:
    # スクリプトと同じディレクトリを検索
    script_dir = Path(__file__).parent
    for f in script_dir.glob("*.xlsx"):
        return f
    # 見つからなければカレントディレクトリも検索
    for f in Path(".").glob("*.xlsx"):
        return f
    raise FileNotFoundError(f"xlsxファイルが見つかりません。検索場所: {script_dir.absolute()} または {Path('.').absolute()}")


def inspect_cell(cell):
    value = cell.value
    return {
        "cell": cell.coordinate,
        "value": value,
        "blank_type": classify_blank(value),  # ★ 追加
        "data_type": cell.data_type,
        "font_name": cell.font.name if cell.font else None,
        "font_color": getattr(cell.font.color, "rgb", None) if cell.font else None,
        "fill_fgcolor": getattr(cell.fill.fgColor, "rgb", None),
        "fill_bgcolor": getattr(cell.fill.bgColor, "rgb", None),
        "number_format": cell.number_format,
    }


def main():
    excel_path = find_excel_file()
    print(f"対象ファイル: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    # 最初のシートを取得（activeの代わり）
    sheet_names = wb.sheetnames
    if not sheet_names:
        raise ValueError("シートが見つかりません")
    ws = wb[sheet_names[0]]
    print(f"シート名: {sheet_names[0]}")

    print(f"=== 列 {TARGET_COLUMN} / 行 {START_ROW}–{END_ROW} を検査 ===")

    for row in range(START_ROW, END_ROW + 1):
        cell = ws[f"{TARGET_COLUMN}{row}"]
        info = inspect_cell(cell)
        print(info)

    print("=== 以上 ===")


if __name__ == "__main__":
    main()
